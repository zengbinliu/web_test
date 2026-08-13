#!/usr/bin/env python3
"""Write gen_tc JSON test cases to Excel (.xlsx)."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:  # pragma: no cover
    raise SystemExit("Missing dependency: pip install openpyxl") from exc

try:
    from validate_cases import format_report, validate_cases
except ImportError:  # pragma: no cover
    validate_cases = None  # type: ignore[assignment]
    format_report = None  # type: ignore[assignment]


COLUMNS = [
    ("id", "用例编号"),
    ("project", "所属项目"),
    ("module", "所属模块"),
    ("title", "用例名称"),
    ("precondition", "前置条件"),
    ("priority", "优先级"),
    ("steps", "操作步骤"),
    ("expects", "预期结果"),
    ("type", "用例类型"),
    ("remark", "备注"),
    ("requirement_ref", "需求来源"),
]

STEP_PREFIX_RE = re.compile(r"^\s*\d+[）)、.．]\s*")


def _normalize_lines(items: list[str]) -> list[str]:
    lines: list[str] = []
    for idx, raw in enumerate(items, start=1):
        text = str(raw).strip()
        if not text:
            continue
        if STEP_PREFIX_RE.match(text):
            lines.append(text)
        else:
            lines.append(f"{idx}）{text}")
    return lines


def _join_lines(items: list[str]) -> str:
    return "\n".join(_normalize_lines(items))


def _auto_id(index: int) -> str:
    return f"TC-{index:03d}"


def _build_title(case: dict[str, Any], meta: dict[str, Any]) -> str:
    title = str(case.get("title") or "").strip()
    version_tag = str(meta.get("version_tag") or "").strip()
    module = str(case.get("module") or "").strip()

    if version_tag and not title.startswith("【"):
        if module and module not in title:
            return f"{version_tag}{module}，{title}"
        return f"{version_tag}{title}"
    if module and module not in title and "，" not in title:
        return f"{module}，{title}"
    return title


def _validate_case(case: dict[str, Any], index: int) -> None:
    steps = case.get("steps") or []
    expects = case.get("expects") or []
    if not isinstance(steps, list) or not isinstance(expects, list):
        raise ValueError(f"case #{index}: steps/expects must be arrays")
    if not steps or not expects:
        raise ValueError(f"case #{index}: steps and expects are required")
    if len(steps) != len(expects):
        raise ValueError(
            f"case #{index}: steps count ({len(steps)}) != expects count ({len(expects)})"
        )
    for field in ("title", "precondition", "priority"):
        if not str(case.get(field) or "").strip():
            raise ValueError(f"case #{index}: missing required field '{field}'")


def load_cases(path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, list):
        return {}, data
    meta = data.get("meta") or {}
    cases = data.get("cases") or []
    if not isinstance(cases, list) or not cases:
        raise ValueError("JSON must contain non-empty 'cases' array")
    return meta, cases


def write_excel(meta: dict[str, Any], cases: list[dict[str, Any]], output: Path, sheet_name: str) -> None:
    for i, case in enumerate(cases, start=1):
        _validate_case(case, i)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [label for _, label in COLUMNS]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    project_default = str(meta.get("project") or "").strip()

    for row_idx, case in enumerate(cases, start=1):
        case_id = str(case.get("id") or _auto_id(row_idx)).strip()
        row = {
            "id": case_id,
            "project": str(case.get("project") or project_default).strip(),
            "module": str(case.get("module") or "").strip(),
            "title": _build_title(case, meta),
            "precondition": str(case.get("precondition") or "").strip(),
            "priority": str(case.get("priority") or "").strip(),
            "steps": _join_lines(case.get("steps") or []),
            "expects": _join_lines(case.get("expects") or []),
            "type": str(case.get("type") or "功能测试").strip(),
            "remark": str(case.get("remark") or "").strip(),
            "requirement_ref": str(case.get("requirement_ref") or "").strip(),
        }
        ws.append([row[key] for key, _ in COLUMNS])

    widths = {
        "A": 12,
        "B": 14,
        "C": 28,
        "D": 48,
        "E": 32,
        "F": 8,
        "G": 44,
        "H": 44,
        "I": 12,
        "J": 24,
        "K": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(COLUMNS)):
        for cell in row:
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Write gen_tc JSON test cases to Excel")
    parser.add_argument("--input", "-i", required=True, help="Input JSON file path")
    parser.add_argument("--output", "-o", required=True, help="Output .xlsx path")
    parser.add_argument("--sheet-name", default="测试用例", help="Worksheet name")
    parser.add_argument(
        "--no-validate",
        action="store_true",
        help="Skip validate_cases.py checks before writing",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail on validation warnings (requires validate_cases.py)",
    )
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()

    if not input_path.is_file():
        print(f"Input not found: {input_path}", file=sys.stderr)
        return 1

    meta, cases = load_cases(input_path)
    sheet_name = str(args.sheet_name or meta.get("sheet_name") or "测试用例")

    if not args.no_validate:
        if validate_cases is None or format_report is None:
            print("Warning: validate_cases.py not found; skipping validation", file=sys.stderr)
        else:
            report = validate_cases(cases)
            print(format_report(report))
            if report.error_count or (args.strict and report.warning_count):
                print("Excel write aborted due to validation issues.", file=sys.stderr)
                return 1

    write_excel(meta, cases, output_path, sheet_name)
    print(f"Wrote {len(cases)} cases -> {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
